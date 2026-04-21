"""Generate a visual Excel file of the TwitchMetrics Prisma schema."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/sebb/Documents/Development/Projects/twitchmetrics/docs/twitchmetrics-schema.xlsx"

# ----------------------------------------------------------
# Schema data — derived from packages/database/prisma/schema.prisma
# ----------------------------------------------------------

DOMAINS = {
    "Auth":      {"color": "1F6FEB", "tables": ["User", "Account", "Session", "VerificationToken"]},
    "Creator":   {"color": "E32C19", "tables": ["CreatorProfile", "CreatorAnalytics", "PlatformAccount",
                                                "MetricSnapshot", "CreatorGrowthRollup"]},
    "Claims":    {"color": "D97706", "tables": ["ClaimRequest"]},
    "Access":    {"color": "7C3AED", "tables": ["TalentManagerAccess"]},
    "Brand":     {"color": "0EA5E9", "tables": ["BrandPartnership"]},
    "Games":     {"color": "10B981", "tables": ["Game", "GameBroadcastLanguage", "GameTopChannel",
                                                "GameClip", "GameViewerSnapshot"]},
    "Leads":     {"color": "EC4899", "tables": ["ReportLead"]},
    "Ops":       {"color": "6B7280", "tables": ["IngestionRun", "AuditLog"]},
}

def domain_of(table):
    for name, d in DOMAINS.items():
        if table in d["tables"]:
            return name, d["color"]
    return "Other", "9CA3AF"

# (table, field, type, nullable, default, is_pk, is_fk, references, on_delete)
FIELDS = [
    # User
    ("User", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("User", "email", "String", True, "", False, False, "", ""),
    ("User", "passwordHash", "String", True, "", False, False, "", ""),
    ("User", "emailVerified", "DateTime", True, "", False, False, "", ""),
    ("User", "name", "String", True, "", False, False, "", ""),
    ("User", "image", "String", True, "", False, False, "", ""),
    ("User", "role", "UserRole", False, "creator", False, False, "", ""),
    ("User", "suspended", "Boolean", False, "false", False, False, "", ""),
    ("User", "hasCompletedOnboarding", "Boolean", False, "false", False, False, "", ""),
    ("User", "createdAt", "DateTime", False, "now()", False, False, "", ""),
    ("User", "updatedAt", "DateTime", False, "@updatedAt", False, False, "", ""),

    # Account
    ("Account", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("Account", "userId", "String (uuid)", False, "", False, True, "User.id", "Cascade"),
    ("Account", "type", "String", False, "", False, False, "", ""),
    ("Account", "provider", "String", False, "", False, False, "", ""),
    ("Account", "providerAccountId", "String", False, "", False, False, "", ""),
    ("Account", "refresh_token", "String", True, "", False, False, "", ""),
    ("Account", "access_token", "String", True, "", False, False, "", ""),
    ("Account", "expires_at", "Int", True, "", False, False, "", ""),
    ("Account", "token_type", "String", True, "", False, False, "", ""),
    ("Account", "scope", "String", True, "", False, False, "", ""),
    ("Account", "id_token", "String", True, "", False, False, "", ""),
    ("Account", "session_state", "String", True, "", False, False, "", ""),

    # Session
    ("Session", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("Session", "sessionToken", "String", False, "", False, False, "", ""),
    ("Session", "userId", "String (uuid)", False, "", False, True, "User.id", "Cascade"),
    ("Session", "expires", "DateTime", False, "", False, False, "", ""),

    # VerificationToken
    ("VerificationToken", "identifier", "String", False, "", False, False, "", ""),
    ("VerificationToken", "token", "String", False, "", False, False, "", ""),
    ("VerificationToken", "expires", "DateTime", False, "", False, False, "", ""),

    # CreatorProfile
    ("CreatorProfile", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("CreatorProfile", "userId", "String (uuid)", True, "", False, True, "User.id", ""),
    ("CreatorProfile", "state", "ProfileState", False, "unclaimed", False, False, "", ""),
    ("CreatorProfile", "snapshotTier", "SnapshotTier", False, "tier3", False, False, "", ""),
    ("CreatorProfile", "displayName", "String", False, "", False, False, "", ""),
    ("CreatorProfile", "slug", "String", False, "", False, False, "", ""),
    ("CreatorProfile", "avatarUrl", "String", True, "", False, False, "", ""),
    ("CreatorProfile", "bannerUrl", "String", True, "", False, False, "", ""),
    ("CreatorProfile", "bio", "String", True, "", False, False, "", ""),
    ("CreatorProfile", "country", "String", True, "", False, False, "", ""),
    ("CreatorProfile", "gender", "String", True, "", False, False, "", ""),
    ("CreatorProfile", "language", "String", True, "", False, False, "", ""),
    ("CreatorProfile", "age", "Int", True, "", False, False, "", ""),
    ("CreatorProfile", "interests", "Json", False, "[]", False, False, "", ""),
    ("CreatorProfile", "primaryPlatform", "Platform", False, "", False, False, "", ""),
    ("CreatorProfile", "totalFollowers", "BigInt", False, "0", False, False, "", ""),
    ("CreatorProfile", "totalViews", "BigInt", False, "0", False, False, "", ""),
    ("CreatorProfile", "searchText", "String", True, '""', False, False, "", ""),
    ("CreatorProfile", "widgetConfig", "Json", False, "[]", False, False, "", ""),
    ("CreatorProfile", "createdAt", "DateTime", False, "now()", False, False, "", ""),
    ("CreatorProfile", "updatedAt", "DateTime", False, "@updatedAt", False, False, "", ""),
    ("CreatorProfile", "lastSnapshotAt", "DateTime", True, "", False, False, "", ""),
    ("CreatorProfile", "lastLinkDiscoveryAt", "DateTime", True, "", False, False, "", ""),
    ("CreatorProfile", "claimedAt", "DateTime", True, "", False, False, "", ""),

    # CreatorAnalytics
    ("CreatorAnalytics", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("CreatorAnalytics", "creatorProfileId", "String (uuid)", False, "", False, True, "CreatorProfile.id", "Cascade"),
    ("CreatorAnalytics", "platform", "Platform", False, "", False, False, "", ""),
    ("CreatorAnalytics", "periodStart", "Date", False, "", False, False, "", ""),
    ("CreatorAnalytics", "periodEnd", "Date", False, "", False, False, "", ""),
    ("CreatorAnalytics", "estimatedMinutesWatched", "BigInt", True, "", False, False, "", ""),
    ("CreatorAnalytics", "averageViewDuration", "Float", True, "", False, False, "", ""),
    ("CreatorAnalytics", "subscribersGained", "Int", True, "", False, False, "", ""),
    ("CreatorAnalytics", "subscribersLost", "Int", True, "", False, False, "", ""),
    ("CreatorAnalytics", "estimatedRevenue", "Float", True, "", False, False, "", ""),
    ("CreatorAnalytics", "views", "BigInt", True, "", False, False, "", ""),
    ("CreatorAnalytics", "likes", "Int", True, "", False, False, "", ""),
    ("CreatorAnalytics", "comments", "Int", True, "", False, False, "", ""),
    ("CreatorAnalytics", "shares", "Int", True, "", False, False, "", ""),
    ("CreatorAnalytics", "impressions", "BigInt", True, "", False, False, "", ""),
    ("CreatorAnalytics", "reach", "BigInt", True, "", False, False, "", ""),
    ("CreatorAnalytics", "profileViews", "Int", True, "", False, False, "", ""),
    ("CreatorAnalytics", "websiteClicks", "Int", True, "", False, False, "", ""),
    ("CreatorAnalytics", "subscriberCount", "Int", True, "", False, False, "", ""),
    ("CreatorAnalytics", "subscriberPoints", "Int", True, "", False, False, "", ""),
    ("CreatorAnalytics", "ageGenderData", "Json", True, "", False, False, "", ""),
    ("CreatorAnalytics", "countryData", "Json", True, "", False, False, "", ""),
    ("CreatorAnalytics", "deviceData", "Json", True, "", False, False, "", ""),
    ("CreatorAnalytics", "trafficSources", "Json", True, "", False, False, "", ""),
    ("CreatorAnalytics", "fetchedAt", "DateTime", False, "now()", False, False, "", ""),
    ("CreatorAnalytics", "createdAt", "DateTime", False, "now()", False, False, "", ""),
    ("CreatorAnalytics", "updatedAt", "DateTime", False, "@updatedAt", False, False, "", ""),

    # PlatformAccount
    ("PlatformAccount", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("PlatformAccount", "creatorProfileId", "String (uuid)", False, "", False, True, "CreatorProfile.id", "Cascade"),
    ("PlatformAccount", "platform", "Platform", False, "", False, False, "", ""),
    ("PlatformAccount", "platformUserId", "String", False, "", False, False, "", ""),
    ("PlatformAccount", "platformUsername", "String", False, "", False, False, "", ""),
    ("PlatformAccount", "platformDisplayName", "String", True, "", False, False, "", ""),
    ("PlatformAccount", "platformUrl", "String", True, "", False, False, "", ""),
    ("PlatformAccount", "platformAvatarUrl", "String", True, "", False, False, "", ""),
    ("PlatformAccount", "accessToken", "String", True, "", False, False, "", ""),
    ("PlatformAccount", "refreshToken", "String", True, "", False, False, "", ""),
    ("PlatformAccount", "tokenExpiresAt", "DateTime", True, "", False, False, "", ""),
    ("PlatformAccount", "oauthScopes", "String[]", False, "[]", False, False, "", ""),
    ("PlatformAccount", "isOAuthConnected", "Boolean", False, "false", False, False, "", ""),
    ("PlatformAccount", "lastOAuthRefresh", "DateTime", True, "", False, False, "", ""),
    ("PlatformAccount", "followerCount", "BigInt", True, "", False, False, "", ""),
    ("PlatformAccount", "followingCount", "BigInt", True, "", False, False, "", ""),
    ("PlatformAccount", "totalViews", "BigInt", True, "", False, False, "", ""),
    ("PlatformAccount", "subscriberCount", "BigInt", True, "", False, False, "", ""),
    ("PlatformAccount", "postCount", "Int", True, "", False, False, "", ""),
    ("PlatformAccount", "lastSyncedAt", "DateTime", True, "", False, False, "", ""),
    ("PlatformAccount", "createdAt", "DateTime", False, "now()", False, False, "", ""),
    ("PlatformAccount", "updatedAt", "DateTime", False, "@updatedAt", False, False, "", ""),

    # MetricSnapshot
    ("MetricSnapshot", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("MetricSnapshot", "creatorProfileId", "String (uuid)", False, "", False, True, "CreatorProfile.id", "Cascade"),
    ("MetricSnapshot", "platform", "Platform", False, "", False, False, "", ""),
    ("MetricSnapshot", "snapshotAt", "DateTime", False, "now()", False, False, "", ""),
    ("MetricSnapshot", "followerCount", "BigInt", True, "", False, False, "", ""),
    ("MetricSnapshot", "followingCount", "BigInt", True, "", False, False, "", ""),
    ("MetricSnapshot", "totalViews", "BigInt", True, "", False, False, "", ""),
    ("MetricSnapshot", "subscriberCount", "BigInt", True, "", False, False, "", ""),
    ("MetricSnapshot", "postCount", "Int", True, "", False, False, "", ""),
    ("MetricSnapshot", "extendedMetrics", "Json", True, "", False, False, "", ""),

    # ClaimRequest
    ("ClaimRequest", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("ClaimRequest", "creatorProfileId", "String (uuid)", False, "", False, True, "CreatorProfile.id", ""),
    ("ClaimRequest", "userId", "String (uuid)", False, "", False, True, "User.id", ""),
    ("ClaimRequest", "method", "ClaimMethod", False, "", False, False, "", ""),
    ("ClaimRequest", "status", "ClaimStatus", False, "pending", False, False, "", ""),
    ("ClaimRequest", "platform", "Platform", False, "", False, False, "", ""),
    ("ClaimRequest", "challengeCode", "String", True, "", False, False, "", ""),
    ("ClaimRequest", "challengeExpiresAt", "DateTime", True, "", False, False, "", ""),
    ("ClaimRequest", "evidenceUrls", "String[]", False, "[]", False, False, "", ""),
    ("ClaimRequest", "reviewNotes", "String", True, "", False, False, "", ""),
    ("ClaimRequest", "reviewedBy", "String (uuid)", True, "", False, False, "", ""),
    ("ClaimRequest", "reviewedAt", "DateTime", True, "", False, False, "", ""),
    ("ClaimRequest", "createdAt", "DateTime", False, "now()", False, False, "", ""),
    ("ClaimRequest", "updatedAt", "DateTime", False, "@updatedAt", False, False, "", ""),
    ("ClaimRequest", "resolvedAt", "DateTime", True, "", False, False, "", ""),

    # TalentManagerAccess
    ("TalentManagerAccess", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("TalentManagerAccess", "managerId", "String (uuid)", False, "", False, True, "User.id", ""),
    ("TalentManagerAccess", "creatorProfileId", "String (uuid)", False, "", False, True, "CreatorProfile.id", ""),
    ("TalentManagerAccess", "canViewAnalytics", "Boolean", False, "true", False, False, "", ""),
    ("TalentManagerAccess", "canEditProfile", "Boolean", False, "false", False, False, "", ""),
    ("TalentManagerAccess", "canExportData", "Boolean", False, "false", False, False, "", ""),
    ("TalentManagerAccess", "canManageBrands", "Boolean", False, "false", False, False, "", ""),
    ("TalentManagerAccess", "grantedAt", "DateTime", False, "now()", False, False, "", ""),
    ("TalentManagerAccess", "grantedBy", "String (uuid)", False, "", False, False, "", ""),
    ("TalentManagerAccess", "revokedAt", "DateTime", True, "", False, False, "", ""),

    # BrandPartnership
    ("BrandPartnership", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("BrandPartnership", "creatorProfileId", "String (uuid)", False, "", False, True, "CreatorProfile.id", "Cascade"),
    ("BrandPartnership", "brandName", "String", False, "", False, False, "", ""),
    ("BrandPartnership", "brandLogoUrl", "String", True, "", False, False, "", ""),
    ("BrandPartnership", "campaignName", "String", True, "", False, False, "", ""),
    ("BrandPartnership", "startDate", "DateTime", True, "", False, False, "", ""),
    ("BrandPartnership", "endDate", "DateTime", True, "", False, False, "", ""),
    ("BrandPartnership", "isPublic", "Boolean", False, "true", False, False, "", ""),
    ("BrandPartnership", "createdAt", "DateTime", False, "now()", False, False, "", ""),

    # Game
    ("Game", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("Game", "twitchGameId", "String", True, "", False, False, "", ""),
    ("Game", "igdbId", "Int", True, "", False, False, "", ""),
    ("Game", "youtubeGameId", "String", True, "", False, False, "", ""),
    ("Game", "name", "String", False, "", False, False, "", ""),
    ("Game", "slug", "String", False, "", False, False, "", ""),
    ("Game", "coverImageUrl", "String", True, "", False, False, "", ""),
    ("Game", "summary", "String", True, "", False, False, "", ""),
    ("Game", "releaseDate", "DateTime", True, "", False, False, "", ""),
    ("Game", "genres", "String[]", False, "[]", False, False, "", ""),
    ("Game", "platforms", "String[]", False, "[]", False, False, "", ""),
    ("Game", "developer", "String", True, "", False, False, "", ""),
    ("Game", "publisher", "String", True, "", False, False, "", ""),
    ("Game", "searchText", "String", True, '""', False, False, "", ""),
    ("Game", "currentViewers", "Int", False, "0", False, False, "", ""),
    ("Game", "currentChannels", "Int", False, "0", False, False, "", ""),
    ("Game", "peakViewers24h", "Int", False, "0", False, False, "", ""),
    ("Game", "avgViewers7d", "Int", False, "0", False, False, "", ""),
    ("Game", "avgLiveChannels", "Int", False, "0", False, False, "", ""),
    ("Game", "hoursWatched7d", "BigInt", False, "0", False, False, "", ""),
    ("Game", "createdAt", "DateTime", False, "now()", False, False, "", ""),
    ("Game", "updatedAt", "DateTime", False, "@updatedAt", False, False, "", ""),

    # GameBroadcastLanguage
    ("GameBroadcastLanguage", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("GameBroadcastLanguage", "gameId", "String (uuid)", False, "", False, True, "Game.id", "Cascade"),
    ("GameBroadcastLanguage", "language", "String", False, "", False, False, "", ""),
    ("GameBroadcastLanguage", "label", "String", False, "", False, False, "", ""),
    ("GameBroadcastLanguage", "percent", "Float", False, "", False, False, "", ""),
    ("GameBroadcastLanguage", "updatedAt", "DateTime", False, "now()", False, False, "", ""),

    # GameTopChannel
    ("GameTopChannel", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("GameTopChannel", "gameId", "String (uuid)", False, "", False, True, "Game.id", "Cascade"),
    ("GameTopChannel", "channelName", "String", False, "", False, False, "", ""),
    ("GameTopChannel", "avatarUrl", "String", True, "", False, False, "", ""),
    ("GameTopChannel", "slug", "String", True, "", False, False, "", ""),
    ("GameTopChannel", "category", "String", False, "", False, False, "", ""),
    ("GameTopChannel", "avgViewers", "Int", False, "0", False, False, "", ""),
    ("GameTopChannel", "airtime", "Int", False, "0", False, False, "", ""),
    ("GameTopChannel", "viewerHours", "BigInt", False, "0", False, False, "", ""),
    ("GameTopChannel", "updatedAt", "DateTime", False, "now()", False, False, "", ""),

    # GameClip
    ("GameClip", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("GameClip", "gameId", "String (uuid)", False, "", False, True, "Game.id", "Cascade"),
    ("GameClip", "clipId", "String", False, "", False, False, "", ""),
    ("GameClip", "title", "String", False, "", False, False, "", ""),
    ("GameClip", "thumbnailUrl", "String", False, "", False, False, "", ""),
    ("GameClip", "url", "String", False, "", False, False, "", ""),
    ("GameClip", "viewCount", "Int", False, "0", False, False, "", ""),
    ("GameClip", "createdAt", "DateTime", False, "", False, False, "", ""),
    ("GameClip", "updatedAt", "DateTime", False, "now()", False, False, "", ""),

    # GameViewerSnapshot
    ("GameViewerSnapshot", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("GameViewerSnapshot", "gameId", "String (uuid)", False, "", False, True, "Game.id", "Cascade"),
    ("GameViewerSnapshot", "snapshotAt", "DateTime", False, "now()", False, False, "", ""),
    ("GameViewerSnapshot", "twitchViewers", "Int", False, "0", False, False, "", ""),
    ("GameViewerSnapshot", "twitchChannels", "Int", False, "0", False, False, "", ""),
    ("GameViewerSnapshot", "youtubeViewers", "Int", False, "0", False, False, "", ""),
    ("GameViewerSnapshot", "youtubeChannels", "Int", False, "0", False, False, "", ""),
    ("GameViewerSnapshot", "kickViewers", "Int", False, "0", False, False, "", ""),
    ("GameViewerSnapshot", "kickChannels", "Int", False, "0", False, False, "", ""),
    ("GameViewerSnapshot", "totalViewers", "Int", False, "0", False, False, "", ""),
    ("GameViewerSnapshot", "totalChannels", "Int", False, "0", False, False, "", ""),

    # ReportLead
    ("ReportLead", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("ReportLead", "name", "String", False, "", False, False, "", ""),
    ("ReportLead", "email", "String", False, "", False, False, "", ""),
    ("ReportLead", "company", "String", False, "", False, False, "", ""),
    ("ReportLead", "details", "Json", True, "", False, False, "", ""),
    ("ReportLead", "createdAt", "DateTime", False, "now()", False, False, "", ""),

    # CreatorGrowthRollup
    ("CreatorGrowthRollup", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("CreatorGrowthRollup", "creatorProfileId", "String (uuid)", False, "", False, True, "CreatorProfile.id", "Cascade"),
    ("CreatorGrowthRollup", "platform", "Platform", False, "", False, False, "", ""),
    ("CreatorGrowthRollup", "followerCount", "BigInt", False, "0", False, False, "", ""),
    ("CreatorGrowthRollup", "delta1d", "BigInt", False, "0", False, False, "", ""),
    ("CreatorGrowthRollup", "delta7d", "BigInt", False, "0", False, False, "", ""),
    ("CreatorGrowthRollup", "delta30d", "BigInt", False, "0", False, False, "", ""),
    ("CreatorGrowthRollup", "pct1d", "Float", False, "0", False, False, "", ""),
    ("CreatorGrowthRollup", "pct7d", "Float", False, "0", False, False, "", ""),
    ("CreatorGrowthRollup", "pct30d", "Float", False, "0", False, False, "", ""),
    ("CreatorGrowthRollup", "trendDirection", "String", False, "FLAT", False, False, "", ""),
    ("CreatorGrowthRollup", "acceleration", "String", False, "STABLE", False, False, "", ""),
    ("CreatorGrowthRollup", "computedAt", "DateTime", False, "now()", False, False, "", ""),
    ("CreatorGrowthRollup", "updatedAt", "DateTime", False, "@updatedAt", False, False, "", ""),

    # IngestionRun
    ("IngestionRun", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("IngestionRun", "domain", "String", False, "", False, False, "", ""),
    ("IngestionRun", "scope", "String", False, "", False, False, "", ""),
    ("IngestionRun", "jobType", "String", False, "", False, False, "", ""),
    ("IngestionRun", "platform", "String", True, "", False, False, "", ""),
    ("IngestionRun", "status", "String", False, "", False, False, "", ""),
    ("IngestionRun", "startedAt", "DateTime", False, "now()", False, False, "", ""),
    ("IngestionRun", "finishedAt", "DateTime", True, "", False, False, "", ""),
    ("IngestionRun", "recordsScanned", "Int", False, "0", False, False, "", ""),
    ("IngestionRun", "recordsWritten", "Int", False, "0", False, False, "", ""),
    ("IngestionRun", "recordsSkipped", "Int", False, "0", False, False, "", ""),
    ("IngestionRun", "recordsFailed", "Int", False, "0", False, False, "", ""),
    ("IngestionRun", "partialCount", "Int", False, "0", False, False, "", ""),
    ("IngestionRun", "quotaUsed", "Int", True, "", False, False, "", ""),
    ("IngestionRun", "errorSummary", "String", True, "", False, False, "", ""),
    ("IngestionRun", "metadata", "Json", True, "", False, False, "", ""),
    ("IngestionRun", "createdAt", "DateTime", False, "now()", False, False, "", ""),

    # AuditLog
    ("AuditLog", "id", "String (uuid)", False, "gen_random_uuid()", True, False, "", ""),
    ("AuditLog", "userId", "String (uuid)", False, "", False, True, "User.id", "Cascade"),
    ("AuditLog", "action", "String", False, "", False, False, "", ""),
    ("AuditLog", "targetType", "String", True, "", False, False, "", ""),
    ("AuditLog", "targetId", "String (uuid)", True, "", False, False, "", ""),
    ("AuditLog", "metadata", "Json", True, "", False, False, "", ""),
    ("AuditLog", "ipAddress", "String", True, "", False, False, "", ""),
    ("AuditLog", "userAgent", "String", True, "", False, False, "", ""),
    ("AuditLog", "createdAt", "DateTime", False, "now()", False, False, "", ""),
]

# Derive relationships from FIELDS
RELATIONSHIPS = []
for f in FIELDS:
    table, field, ftype, nullable, default, is_pk, is_fk, references, on_delete = f
    if is_fk and references:
        tgt_table, tgt_field = references.split(".")
        RELATIONSHIPS.append({
            "from_table": table, "from_field": field,
            "to_table": tgt_table, "to_field": tgt_field,
            "nullable": "nullable" if nullable else "required",
            "on_delete": on_delete or "—",
        })

ENUMS = {
    "UserRole":     ["creator", "talent_manager", "admin"],
    "ProfileState": ["unclaimed", "pending_claim", "claimed", "premium"],
    "Platform":     ["twitch", "youtube", "instagram", "tiktok", "x", "kick"],
    "SnapshotTier": ["tier1", "tier2", "tier3"],
    "ClaimMethod":  ["oauth", "cross_platform", "bio_challenge", "post_challenge", "manual_review"],
    "ClaimStatus":  ["pending", "approved", "rejected", "expired"],
}

# Table descriptions
TABLE_DESC = {
    "User": "Auth.js user — creators, talent managers, admins.",
    "Account": "Auth.js OAuth/credentials accounts linked to a User.",
    "Session": "Active Auth.js sessions.",
    "VerificationToken": "Auth.js email verification tokens.",
    "CreatorProfile": "Canonical creator record. Public profile, aggregate metrics, widget config.",
    "CreatorAnalytics": "Per-platform per-period analytics (watch time, revenue, subs, demographics).",
    "PlatformAccount": "Connected external platform account (Twitch/YouTube/etc) for a creator.",
    "MetricSnapshot": "Time-series follower/view counts per platform — powers charts.",
    "CreatorGrowthRollup": "Precomputed 1d/7d/30d deltas & trend direction per creator/platform.",
    "ClaimRequest": "Request from a User to claim a CreatorProfile (oauth/bio/post challenges).",
    "TalentManagerAccess": "Permission grant for a manager User to view/edit a CreatorProfile.",
    "BrandPartnership": "Past or current brand sponsorship for a creator.",
    "Game": "Canonical game/category record (Twitch + IGDB + YouTube IDs).",
    "GameBroadcastLanguage": "Language distribution of live streams for a game.",
    "GameTopChannel": "Top-streaming channels per game (denormalized).",
    "GameClip": "Featured clips for a game.",
    "GameViewerSnapshot": "Time-series per-platform viewer/channel counts for a game.",
    "ReportLead": "Marketing lead from the /reports lead-gen form.",
    "IngestionRun": "Observability record for each worker/cron ingestion run.",
    "AuditLog": "Admin/security audit trail for sensitive actions.",
}

# ----------------------------------------------------------
# Styling helpers
# ----------------------------------------------------------
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="111827")
DOMAIN_LABEL_FONT = Font(bold=True, size=12, color="111827")
PK_FILL = PatternFill("solid", fgColor="FEF3C7")
FK_FILL = PatternFill("solid", fgColor="DBEAFE")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ----------------------------------------------------------
# Build workbook
# ----------------------------------------------------------
wb = Workbook()

# -----------------------------
# Sheet 1: README
# -----------------------------
ws = wb.active
ws.title = "README"
ws["A1"] = "TwitchMetrics — Database Schema"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Source: packages/database/prisma/schema.prisma  •  PostgreSQL (Neon) via Prisma 6.19"
ws["A2"].font = Font(italic=True, color="6B7280")

ws["A4"] = "Sheets"
ws["A4"].font = DOMAIN_LABEL_FONT
sheet_info = [
    ("Schema Map",    "Visual layout of tables grouped by domain — each card shows the table header + its fields."),
    ("Tables",        "Every table with its domain, description, field count, and FK in/out counts."),
    ("Fields",        "Every column across every table. PKs highlighted amber, FKs highlighted blue."),
    ("Relationships", "Every foreign key edge (from → to) with cascade rule and nullability."),
    ("Enums",         "All Prisma enums with allowed values."),
]
ws["A5"] = "Sheet"; ws["B5"] = "Purpose"
ws["A5"].font = HEADER_FONT; ws["B5"].font = HEADER_FONT
ws["A5"].fill = header_fill("111827"); ws["B5"].fill = header_fill("111827")
for i, (name, desc) in enumerate(sheet_info, start=6):
    ws.cell(row=i, column=1, value=name).alignment = LEFT
    ws.cell(row=i, column=2, value=desc).alignment = LEFT
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=2).border = BORDER

ws["A13"] = "Domain Legend"
ws["A13"].font = DOMAIN_LABEL_FONT
ws["A14"] = "Domain"; ws["B14"] = "Tables"
ws["A14"].font = HEADER_FONT; ws["B14"].font = HEADER_FONT
ws["A14"].fill = header_fill("111827"); ws["B14"].fill = header_fill("111827")
r = 15
for dname, d in DOMAINS.items():
    ca = ws.cell(row=r, column=1, value=dname)
    cb = ws.cell(row=r, column=2, value=", ".join(d["tables"]))
    ca.fill = header_fill(d["color"])
    ca.font = Font(bold=True, color="FFFFFF")
    ca.alignment = LEFT
    cb.alignment = LEFT
    ca.border = BORDER; cb.border = BORDER
    r += 1

ws[f"A{r+1}"] = "Field Highlighting"
ws[f"A{r+1}"].font = DOMAIN_LABEL_FONT
ws[f"A{r+2}"] = "Primary key (PK)"; ws[f"A{r+2}"].fill = PK_FILL; ws[f"A{r+2}"].border = BORDER
ws[f"A{r+3}"] = "Foreign key (FK)"; ws[f"A{r+3}"].fill = FK_FILL; ws[f"A{r+3}"].border = BORDER

set_col_widths(ws, [22, 90])

# -----------------------------
# Sheet 2: Schema Map — card layout
# -----------------------------
ws = wb.create_sheet("Schema Map")
ws["A1"] = "Schema Map"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Each card is a table. Header color = domain. Blue rows are foreign keys pointing to another table."
ws["A2"].font = Font(italic=True, color="6B7280")

# Cards laid out 3 per row, column widths uniform
CARD_COLS = 3
CARD_WIDTH = 3     # each card spans 3 columns (Field | Type | FK→)
CARD_GAP = 1
CARD_COL_WIDTHS = [22, 18, 22]

tables_in_order = []
for dname, d in DOMAINS.items():
    for t in d["tables"]:
        tables_in_order.append(t)

# set column widths for full sheet based on cards
for card_idx in range(CARD_COLS):
    base = card_idx * (CARD_WIDTH + CARD_GAP) + 1
    for j, w in enumerate(CARD_COL_WIDTHS):
        ws.column_dimensions[get_column_letter(base + j)].width = w
    if card_idx < CARD_COLS - 1:
        ws.column_dimensions[get_column_letter(base + CARD_WIDTH)].width = 2

# group fields by table for quick lookup
from collections import defaultdict
fields_by_table = defaultdict(list)
for f in FIELDS:
    fields_by_table[f[0]].append(f)

row_cursor = 4
row_heights_in_cards = []
card_starts_this_row = []
max_rows_this_row = 0

def write_card(table, start_row, start_col):
    dname, color = domain_of(table)
    end_col = start_col + CARD_WIDTH - 1

    # header: table name + domain tag
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    hc = ws.cell(row=start_row, column=start_col, value=f"{table}   ·   {dname}")
    hc.fill = header_fill(color)
    hc.font = Font(bold=True, color="FFFFFF", size=12)
    hc.alignment = CENTER
    hc.border = BORDER
    ws.row_dimensions[start_row].height = 22

    # column headers
    hdr_row = start_row + 1
    for i, label in enumerate(["Field", "Type", "FK →"]):
        c = ws.cell(row=hdr_row, column=start_col + i, value=label)
        c.fill = header_fill("111827")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = CENTER
        c.border = BORDER

    # fields
    r = hdr_row + 1
    for _, field, ftype, nullable, default, is_pk, is_fk, references, on_delete in fields_by_table[table]:
        name_cell = ws.cell(row=r, column=start_col, value=("🔑 " if is_pk else "") + field + ("" if not nullable else " ?"))
        type_cell = ws.cell(row=r, column=start_col + 1, value=ftype)
        fk_cell   = ws.cell(row=r, column=start_col + 2, value=(references if is_fk else ""))

        for c in (name_cell, type_cell, fk_cell):
            c.border = BORDER
            c.alignment = LEFT
            c.font = Font(size=10)

        if is_pk:
            name_cell.fill = PK_FILL
            type_cell.fill = PK_FILL
        if is_fk:
            name_cell.fill = FK_FILL
            type_cell.fill = FK_FILL
            fk_cell.fill = FK_FILL
            fk_cell.font = Font(size=10, italic=True, color="1E40AF")
        r += 1
    # description footer
    ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=end_col)
    dc = ws.cell(row=r, column=start_col, value=TABLE_DESC.get(table, ""))
    dc.font = Font(size=9, italic=True, color="6B7280")
    dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    dc.border = BORDER
    ws.row_dimensions[r].height = 32
    return r  # last row used

i = 0
while i < len(tables_in_order):
    chunk = tables_in_order[i:i + CARD_COLS]
    max_end = row_cursor
    for idx, table in enumerate(chunk):
        start_col = idx * (CARD_WIDTH + CARD_GAP) + 1
        end_row = write_card(table, row_cursor, start_col)
        if end_row > max_end:
            max_end = end_row
    row_cursor = max_end + 2
    i += CARD_COLS

ws.freeze_panes = "A4"

# -----------------------------
# Sheet 3: Tables
# -----------------------------
ws = wb.create_sheet("Tables")
ws["A1"] = "Tables"
ws["A1"].font = TITLE_FONT
headers = ["Table", "Domain", "Description", "Fields", "FKs Out", "FKs In"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = header_fill("111827")
    c.alignment = CENTER
    c.border = BORDER

fks_in_counts = defaultdict(int)
for rel in RELATIONSHIPS:
    fks_in_counts[rel["to_table"]] += 1

row = 4
for table in tables_in_order:
    dname, color = domain_of(table)
    field_count = len(fields_by_table[table])
    fks_out = sum(1 for f in fields_by_table[table] if f[6])
    fks_in  = fks_in_counts.get(table, 0)

    cells = [
        ws.cell(row=row, column=1, value=table),
        ws.cell(row=row, column=2, value=dname),
        ws.cell(row=row, column=3, value=TABLE_DESC.get(table, "")),
        ws.cell(row=row, column=4, value=field_count),
        ws.cell(row=row, column=5, value=fks_out),
        ws.cell(row=row, column=6, value=fks_in),
    ]
    cells[1].fill = header_fill(color)
    cells[1].font = Font(bold=True, color="FFFFFF")
    cells[1].alignment = CENTER
    for c in cells:
        c.border = BORDER
        if c.column != 2:
            c.alignment = LEFT if c.column in (1, 3) else CENTER
    row += 1

set_col_widths(ws, [26, 14, 75, 10, 10, 10])
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:F{row-1}"

# -----------------------------
# Sheet 4: Fields
# -----------------------------
ws = wb.create_sheet("Fields")
ws["A1"] = "Fields"
ws["A1"].font = TITLE_FONT
headers = ["Table", "Field", "Type", "Nullable", "Default", "PK", "FK", "References", "On Delete"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = header_fill("111827")
    c.alignment = CENTER
    c.border = BORDER

row = 4
for table, field, ftype, nullable, default, is_pk, is_fk, references, on_delete in FIELDS:
    values = [
        table, field, ftype,
        "Yes" if nullable else "No",
        default or "",
        "✓" if is_pk else "",
        "✓" if is_fk else "",
        references or "",
        on_delete or "",
    ]
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        c.alignment = LEFT if i in (1,2,3,5,8,9) else CENTER
        c.font = Font(size=10)
    if is_pk:
        for i in range(1, 10):
            ws.cell(row=row, column=i).fill = PK_FILL
    elif is_fk:
        for i in range(1, 10):
            ws.cell(row=row, column=i).fill = FK_FILL
    row += 1

set_col_widths(ws, [24, 26, 16, 10, 20, 6, 6, 28, 12])
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:I{row-1}"

# -----------------------------
# Sheet 5: Relationships
# -----------------------------
ws = wb.create_sheet("Relationships")
ws["A1"] = "Relationships"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Each row is a foreign-key edge. Arrow column shows connector direction."
ws["A2"].font = Font(italic=True, color="6B7280")

headers = ["From Table", "From Field", "→", "To Table", "To Field", "Nullability", "On Delete"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = header_fill("111827")
    c.alignment = CENTER
    c.border = BORDER

row = 5
for rel in sorted(RELATIONSHIPS, key=lambda r: (r["from_table"], r["from_field"])):
    from_color = domain_of(rel["from_table"])[1]
    to_color   = domain_of(rel["to_table"])[1]
    vals = [
        rel["from_table"], rel["from_field"], "➜",
        rel["to_table"], rel["to_field"], rel["nullable"], rel["on_delete"],
    ]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        c.alignment = CENTER if i in (3, 6, 7) else LEFT
        c.font = Font(size=10)
    ws.cell(row=row, column=1).fill = header_fill(from_color)
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=row, column=4).fill = header_fill(to_color)
    ws.cell(row=row, column=4).font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=row, column=3).font = Font(bold=True, size=14, color="111827")
    if rel["on_delete"] == "Cascade":
        ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor="FEE2E2")
    row += 1

set_col_widths(ws, [22, 22, 4, 22, 14, 14, 12])
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:G{row-1}"

# -----------------------------
# Sheet 6: Enums
# -----------------------------
ws = wb.create_sheet("Enums")
ws["A1"] = "Enums"
ws["A1"].font = TITLE_FONT
ws["A3"] = "Enum"; ws["B3"] = "Values"
ws["A3"].font = HEADER_FONT; ws["B3"].font = HEADER_FONT
ws["A3"].fill = header_fill("111827"); ws["B3"].fill = header_fill("111827")
ws["A3"].alignment = CENTER; ws["B3"].alignment = CENTER
ws["A3"].border = BORDER; ws["B3"].border = BORDER

row = 4
for name, values in ENUMS.items():
    ca = ws.cell(row=row, column=1, value=name)
    cb = ws.cell(row=row, column=2, value=", ".join(values))
    ca.font = Font(bold=True)
    ca.alignment = LEFT; cb.alignment = LEFT
    ca.border = BORDER; cb.border = BORDER
    row += 1

set_col_widths(ws, [20, 80])

# ----------------------------------------------------------
wb.save(OUT)
print(f"Wrote: {OUT}")
print(f"Tables: {len(tables_in_order)}, Fields: {len(FIELDS)}, Relationships: {len(RELATIONSHIPS)}, Enums: {len(ENUMS)}")
